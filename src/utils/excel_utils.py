#-*- encoding: utf-8 -*-


import openpyxl
#workbook相关
from openpyxl.workbook import Workbook
#封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
#一个数字转为列字母的方法
# from openpyxl.cell import get_column_letter


def open_excel(filePath, readOnly = False):
    '''
        打开excel文件
        filePath:excel文件路径
    '''
    try:
        data = openpyxl.load_workbook(filePath, read_only = readOnly)
        return data
    except Exception as e:
        print(str(e))
        return None


def get_active_sheet(wb):
    '''
        返回当前活动的sheet
        wb：excel的workbook对象
    '''
    return wb.active


def get_sheet_by_name(wb, title):
    '''
        返回标题匹配的sheet
        wb：excel的workbook对象
        title：sheet的标题
    '''
    return wb.get_sheet_by_name(title)

def get_sheet_names(wb):
    '''
        返回sheet的标题数组
    '''
    return wb.sheetnames


def get_cell_val(ws, row, col):
    '''
        返回sheet的单元格的值
        ws:sheet对象
        row:行索引，0开始
        col:列索引，0开始
    '''
    return ws.cell(row = row, column = col).value





def create_excel():
    '''
        创建excel文件
    '''
    try:
        wb = Workbook()
        return wb
    except Exception as e:
        print(str(e))
        return None

def create_sheet(wb, title = '', index = None):
    '''
        创建sheet工作表
        wb：excel的workbook对象
        title：工作表标题
        index：工作表索引，从0开始，不传默认加到最后
    '''
    try:
        ws = None
        if None == index:
            ws = wb.create_sheet(title = '')
        else:
            ws = wb.create_sheet(index = index, title = '')
        return ws
    except Exception as e:
        print(str(e))
        return None


def set_sheet_title(ws, title = ''):
    '''
        设置sheet工作表标题
        ws：excel的sheet对象
        title：工作表标题
    '''
    ws.title = title


def set_cell_val(ws, row, col, val):
    '''
        返回sheet的单元格的值
        ws:sheet对象
        row:行索引，0开始
        col:列索引，0开始
        val: 单元格的值
    '''
    c = ws.cell(row = row, column = col)
    c.value = val


def save(filePath, wb):
    '''
        保存excel
        filePath：保存文件路径
        wb：excel的workbook对象
    '''
    wb.save(filePath)



if __name__ == '__main__':
    # wb = create_excel()
    # ws = get_active_sheet(wb)
    # for i in range(0, 10):
    #     for j in range(0, 12):
    #         set_cell_val(ws, i, j, str(i)+'--'+str(j))
    # save('C:\\Users\\Administrator\\Desktop\\a.xlsx', wb)
    wb = open_excel('../../test/excel/demo1.xlsx', True)
    print(get_sheet_names(wb))
    # for i in tuple(get_active_sheet(wb).rows):
    #     print(i)
    # for i in tuple(get_active_sheet(wb).columns):
    #     print(i)

    for row in get_active_sheet(wb).values:
        print("--")
        for value in row:
            print(value)
